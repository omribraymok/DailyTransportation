import math
# Calculation of random points
import random
# import time
from collections import namedtuple
from datetime import datetime

import numpy as np
# Reading and Writing from/to excel file
import openpyxl
import matplotlib.pyplot as plt

from clCar import Car
from clChild import Child
from clSchool import School
from k_means import k_means
from datetime import timedelta

Result = namedtuple('Result', 'clusters means, school_point list_of_address_in_short_path')


def _calculate_cost_time_random(list_of_all_children, k, point_list,
                                time_matrix, list_of_car, school):
    total_time = 0
    total_cost = 0
    random.shuffle(list_of_all_children)
    j = 0
    divide_list_of_children = [list_of_all_children[x::3] for x in range(k)]
    for temp_list_of_children in divide_list_of_children:
        total_time_per_group = 0
        length = len(temp_list_of_children)
        for i in range(length - 1):
            time = list_of_car[j].calculate_time(temp_list_of_children[i].address, temp_list_of_children[i + 1].address,
                                                 point_list, time_matrix)
            total_time_per_group += time
        # calculate the time and cost for school
        time = list_of_car[j].calculate_time(temp_list_of_children[length - 1].address, school.address,
                                             point_list, time_matrix)
        total_time_per_group += time
        cost = 0
        if total_time_per_group > list_of_car[j].tMin:
            cost = (total_time_per_group - list_of_car[j].tMin) * list_of_car[j].cost_per_minute
        total_cost_per_group = list_of_car[j].driver_cost + cost
        total_cost += total_cost_per_group
        total_time += total_time_per_group
        j += 1
    return total_cost, total_time


def divide_list_of_children_by_k_means(k, point_list, list_of_all_children):
    children_coordinates = []
    divide_list_of_children = []
    temp_list_of_children = []

    for i in point_list:
        children_coordinates.append(list(i))
    children_coordinates.remove([0, 0])

    clusters, means = k_means(children_coordinates, k)

    for i in clusters:
        for j in i:
            x1, y1 = j
            index = point_list.index((x1, y1))
            temp = list_of_all_children[index]
            temp_list_of_children.append(temp)

        new_list = temp_list_of_children.copy()
        divide_list_of_children.append(new_list)
        temp_list_of_children.clear()

    return divide_list_of_children, means, clusters


# Python function to get permutations of a given list
# https://www.geeksforgeeks.org/generate-all-the-permutation-of-a-list-in-python/
def _permutation(lst):
    # If lst is empty then there are no permutations
    if len(lst) == 0:
        return []

    # If there is only one element in lst then, only
    # one permutation is possible
    if len(lst) == 1:
        return [lst]

    # Find the permutations for lst if there are
    # more than 1 characters

    l = []  # empty list that will store current permutation

    # Iterate the input(lst) and calculate the permutation
    for i in range(len(lst)):
        m = lst[i]

        # Extract lst[i] or m from the list.  remLst is
        # remaining list
        remLst = lst[:i] + lst[i + 1:]

        # Generating all permutations where m is first
        # element
        for p in _permutation(remLst):
            l.append([m] + p)
    return l


def _calculate_euclidean_dist(start, destination):
    (x_1, y_1) = start
    (x_2, y_2) = destination
    temp_var = (x_1 - x_2) ** 2 + (y_1 - y_2) ** 2
    return math.sqrt(temp_var) * random.uniform(1, 1.5)


# For each group calculate the time and cost
# The function checks all possible path and choosing the shortest path
def _calculate_cost_time_per_group(list_of_children, car, school,
                                   point_list, time_matrix, length_list_of_all_children, group_number):
    total_time = 0
    total_time_temp_perm = 0
    length = len(list_of_children)
    short_path = []
    time_for_each_children_on_short_path = []
    flag = 0
    times_of_all_permutation = []
    if length < 8:
        for temp_list_of_children in _permutation(list_of_children):
            times = []
            addresses = [temp_list_of_children[0].address]
            for i in range(length - 1):
                time = car.calculate_time(temp_list_of_children[i].address, temp_list_of_children[i + 1].address,
                                          point_list, time_matrix)
                total_time_temp_perm += time
                times.append(time)
                addresses.append(temp_list_of_children[i + 1].address)
            # calculate the time and cost for school
            time = car.calculate_time(temp_list_of_children[length - 1].address, school.address, point_list,
                                      time_matrix)
            total_time_temp_perm += time
            addresses.append(temp_list_of_children[length - 1].address)  # twice
            addresses.append(school.address)
            times.append(time)
            cost = 0
            if total_time_temp_perm > car.tMin:
                cost = (total_time_temp_perm - car.tMin) * car.cost_per_minute
            total_cost_temp_perm = car.driver_cost + cost

            times_of_all_permutation.append(total_time_temp_perm)

            if total_time_temp_perm < total_time or flag == 0:
                flag = 1
                short_path = temp_list_of_children
                total_time = total_time_temp_perm
                total_cost = total_cost_temp_perm
                time_for_each_children_on_short_path = times.copy()

            total_time_temp_perm = 0
    else:
        for j in range(10000):
            temp_list_of_children = np.random.permutation(list_of_children)
            times = []
            addresses = [temp_list_of_children[0].address]
            for i in range(length - 1):
                time = car.calculate_time(temp_list_of_children[i].address, temp_list_of_children[i + 1].address,
                                          point_list, time_matrix)
                total_time_temp_perm += time
                times.append(time)
                addresses.append(temp_list_of_children[i + 1].address)
            # calculate the time and cost for school
            time = car.calculate_time(temp_list_of_children[length - 1].address, school.address, point_list,
                                      time_matrix)
            total_time_temp_perm += time
            addresses.append(temp_list_of_children[length - 1].address)  # twice
            addresses.append(school.address)
            times.append(time)
            cost = 0
            if total_time_temp_perm > car.tMin:
                cost = (total_time_temp_perm - car.tMin) * car.cost_per_minute
            total_cost_temp_perm = car.driver_cost + cost
            total_time_temp_perm += time

            times_of_all_permutation.append(total_time_temp_perm)

            if total_time_temp_perm < total_time or flag == 0:
                flag = 1
                short_path = temp_list_of_children
                total_time = total_time_temp_perm
                total_cost = total_cost_temp_perm
                time_for_each_children_on_short_path = times.copy()

            total_time_temp_perm = 0

    # Turn interactive plotting off
    plt.ioff()
    ypos = np.arange(len(times_of_all_permutation))
    plt.ylabel("time")
    plt.xlabel("Routes")
    plt.title("Histogram of group:" + str(group_number))
    barlist = plt.bar(ypos, times_of_all_permutation)
    i = times_of_all_permutation.index(total_time)
    barlist[i].set_color('r')
    plt.title("Histogram of group:" + str(group_number))
    plt.savefig("output/Histogram of group" + str(group_number) + ".jpg")
    plt.cla()

    temp = 0
    length = len(time_for_each_children_on_short_path)
    for i in range(length - 1, -1, -1):
        temp += time_for_each_children_on_short_path[i]
        time_for_each_children_on_short_path[i] = temp

    print_to_excel_time_cost_per_group(short_path, car, total_cost, total_time,
                                       time_for_each_children_on_short_path, time_matrix, school,
                                       length_list_of_all_children, group_number)

    list_of_address_in_short_path = []
    for child in short_path:
        (x, y) = child.address.split(',')
        (x, y) = (float(x), float(y))
        list_of_address_in_short_path.append((x, y))
    (x, y) = school.address.split(',')
    (x, y) = (float(x), float(y))
    list_of_address_in_short_path.append((x, y))

    return list_of_address_in_short_path, total_cost, total_time


def load_data(data_file):
    # number_of_children = 21
    # number_of_cars = 3

    # To open Workbook
    excel_file = openpyxl.load_workbook(data_file)

    # To open sheets
    child_sheet = excel_file.worksheets[0]
    number_of_children = child_sheet.max_row - 3
    list_of_all_children = []

    # get data from excel file
    for number_row in range(2, number_of_children + 2):
        ch = Child(child_sheet, number_row)
        ch.id = number_row - 2
        list_of_all_children.insert((number_row - 2), ch)

    # Reading from excel file from Car table
    cars_sheet = excel_file.worksheets[1]
    number_of_cars = cars_sheet.max_row - 1
    list_of_cars = []
    # get cars data from data sheet
    for number_row in range(2, number_of_cars + 2):
        list_of_cars.insert((number_row - 2), (Car(cars_sheet, number_row)))

    # Reading from excel file from School table
    schools_sheet = excel_file.worksheets[2]

    list_of_school = []
    for number_row in range(2, 3):
        list_of_school.insert((number_row - 2), (School(schools_sheet, number_row)))

    return number_of_children, list_of_all_children, list_of_cars, list_of_school


def calculate(number_of_children, list_of_all_children, list_of_cars, list_of_school, k_count):
    if k_count > len(list_of_cars):
        raise ValueError(f'K ({k_count}) bigger than the amount of defined transports in the '
                         f'data file ({len(list_of_cars)})')

    # Enter the school's address to the matrix
    temp_point = list_of_school[0].address
    (x_1, y_1) = temp_point.split(',')
    (x_1, y_1) = (float(x_1), float(y_1))

    point_list = []
    point_list.insert(0, (x_1, y_1))

    # Enter the children's address to the list
    for number_row in range(0, number_of_children):
        temp_point = list_of_all_children[number_row].address
        (x1, y1) = temp_point.split(',')
        (x1, y1) = (float(x1), float(y1))
        point_list.insert(number_row, (x1, y1))

    # Matrix 22X22, this matrix will contain all the time travel from point A (row0) to point B (column0)
    time_matrix = np.zeros((number_of_children + 1, number_of_children + 1))
    length_list_of_all_children = len(list_of_all_children)

    for i in range(length_list_of_all_children + 1):
        for j in range(length_list_of_all_children + 1):
            time_matrix[i, j] = math.ceil(_calculate_euclidean_dist(point_list[i], point_list[j]))

    divide_list_of_children, means, clusters = divide_list_of_children_by_k_means(k_count,
                                                                                  point_list, list_of_all_children)
    for i, cluster in enumerate(clusters):
        if len(cluster) >= 14:
            raise ValueError(f"resulting cluster {i} is larger than 14")

    length = len(divide_list_of_children)
    list_of_address_in_short_path = [None] * length

    total_cost_k_means = 0
    total_time_k_means = 0
    for i in range(length):
        print('calc time cost', i)
        (list_of_address_in_short_path[i], total_cost_per_group, total_time_per_group) = _calculate_cost_time_per_group(
            divide_list_of_children[i], list_of_cars[i],
            list_of_school[0], point_list, time_matrix,
            length_list_of_all_children, group_number=i)
        total_cost_k_means += total_cost_per_group
        total_time_k_means += total_time_per_group
        print('done', i)

    (random_cost, random_time) = _calculate_cost_time_random(list_of_all_children, k_count, point_list, time_matrix,
                                                             list_of_cars, list_of_school[0])

    label_bar_chart = ['cost', 'time']
    random_result = [random_cost, random_time]
    algorithm_result = [total_cost_k_means, total_time_k_means]
    xpos = np.arange(len(label_bar_chart))
    plt.xticks(xpos, label_bar_chart)
    plt.title("Comparison between algorithm and random")
    plt.bar(xpos - 0.2, random_result, width=0.4, label="random result")
    plt.bar(xpos + 0.2, algorithm_result, width=0.4, label="algorithm result")
    plt.legend()
    plt.savefig("output/Histogram of result.jpg")
    plt.cla()

    return Result(clusters, means, (x_1, y_1), list_of_address_in_short_path)


# Creates an excel file for each group
# that contains the travel path, cost and travel time
def print_to_excel_time_cost_per_group(lst, car, cost, time, time_for_each_children_on_short_path, time_matrix, school,
                                       length_list_of_all_children, group_number, spath="myPath", sschool="mySchool",
                                       timeOfStart=timedelta(hours=7)):
    # create workbook
    wb = openpyxl.Workbook()
    # get worksheet
    ws = wb.active
    # # change sheet name
    # ws.title = "Group" + str(sheet_number)

    irow = 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "path id"
    tab = ws.cell(row=irow, column=2)
    tab.value = spath
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "car.ID"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(car.ID)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "School"
    tab = ws.cell(row=irow, column=2)
    tab.value = sschool
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "nChildren"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(len(lst))
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "timeOfStart"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(timeOfStart)

    currentTime = timeOfStart
    for t in range(len(lst)):
        # get a pointer for tab in table
        if t > 0:
            iInMatrix = lst[t - 1].id
            jInMatrix = lst[t].id
            temp_time = time_matrix[iInMatrix][jInMatrix]
            currentTime = currentTime + timedelta(minutes=temp_time)
        irow = irow + 1
        tab = ws.cell(row=irow, column=1)
        tab.value = str(t)
        tab = ws.cell(row=irow, column=2)
        tab.value = str(lst[t].id)
        tab = ws.cell(row=irow, column=3)
        tab.value = lst[t].first_name
        tab = ws.cell(row=irow, column=4)
        tab.value = lst[t].lest_name
        tab = ws.cell(row=irow, column=5)
        tab.value = lst[t].address
        tab = ws.cell(row=irow, column=6)
        tab.value = lst[t].contacts
        tab = ws.cell(row=irow, column=7)
        tab.value = str(currentTime)
        tab = ws.cell(row=irow, column=8)
        tab.value = str(time_for_each_children_on_short_path[t])
    iInMatrix = lst[len(lst) - 1].id
    jInMatrix = length_list_of_all_children
    temp_time = time_matrix[iInMatrix][jInMatrix]
    currentTime = currentTime + timedelta(minutes=temp_time)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "school"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(school.id)
    tab = ws.cell(row=irow, column=3)
    tab.value = school.name
    tab = ws.cell(row=irow, column=4)
    tab.value = sschool
    tab = ws.cell(row=irow, column=5)
    tab.value = school.address
    tab = ws.cell(row=irow, column=6)
    tab.value = school.contacts
    tab = ws.cell(row=irow, column=7)
    tab.value = str(currentTime)

    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "cost"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(cost)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "time"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(time)
    wb.save("Groups" + str(group_number) + ".xlsx")
