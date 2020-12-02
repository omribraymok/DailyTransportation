# Reading and Writing from/to excel file
import openpyxl
class Human:

    def __init__(self, ID, first_name, lest_name, address, contacts):
        self.id = ID
        self.first_name = first_name
        self.lest_name = lest_name
        self.address = address
        self.contacts = contacts

    def __repr__(self):
        return "('{}', '{}', '{}', '{}')".format(self.id, self.first_name, self.lest_name, self.address)

    def __str__(self):
        return "({}, {} {}: {})".format(self.id, self.first_name, self.lest_name, self.address)


class Child(Human):
    num_of_children = 0

    def __init__(self, ID, first_name, lest_name, address, contacts, id_school, special_needs):
        super().__init__(ID, first_name, lest_name, address, contacts)
        self.id_school = id_school
        self.spacial_needs = special_needs

        Child.num_of_children += 1

    def __init__(self, child_sheet, number_row):
        self.id = child_sheet.cell(row=number_row, column=1)
        self.first_name = child_sheet.cell(row=number_row, column=2)
        self.lest_name = child_sheet.cell(row=number_row, column=3)
        self.address = child_sheet.cell(row=number_row, column=4)
        self.contacts = child_sheet.cell(row=number_row, column=5)
        self.id_school = child_sheet.cell(row=number_row, column=6)
        self.spacial_needs = child_sheet.cell(row=number_row, column=7)

        Child.num_of_children += 1


class Accompanier(Human):
    number_of_accompaniers = 0

    def __init__(self, ID, first_name, lest_name, address, contacts, preferences, history, special):
        super().__init__(ID, first_name, lest_name, address, contacts)
        self.preferences = preferences
        self.history = history
        self.special = special

        Accompanier.number_of_accompaniers += 1
