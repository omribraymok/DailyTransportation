from clCar import Car
from clHuman import Child
from clSchool import School
from k_means import k_means
import global_variables as gv
# Reading and Writing from/to excel file
import openpyxl
# Calculation of random points
import random
import math
import numpy as np
# import time
import matplotlib.pyplot as plt
import networkx as nx
from datetime import datetime, timedelta


def divide_list_of_children_by_k_means(k):
    children_coordinates = []
    divide_list_of_children = []
    temp_list_of_children = []
    for i in gv.point_list:
        children_coordinates.append(list(i))
    children_coordinates.remove([0, 0])
    gv.clusters = k_means(children_coordinates, k)
    for i in gv.clusters:
        for j in i:
            (x1, y1) = j
            index = gv.point_list.index((x1, y1))
            temp = gv.list_of_all_children[index]
            temp_list_of_children.append(temp)
        new_list = temp_list_of_children.copy()
        divide_list_of_children.append(new_list)
        temp_list_of_children.clear()
    return divide_list_of_children


# Python function to get permutations of a given list
# https://www.geeksforgeeks.org/generate-all-the-permutation-of-a-list-in-python/
def permutation(lst):
    # If lst is empty then there are no permutations
    if len(lst) == 0:
        return []

    # If there is only one element in lst then, only
    # one permutation is possible
    if len(lst) == 1:
        return [lst]

    # Find the permutations for lst if there are
    # more than 1 characters

    l = []  # empty list that will store current permutation

    # Iterate the input(lst) and calculate the permutation
    for i in range(len(lst)):
        m = lst[i]

        # Extract lst[i] or m from the list.  remLst is
        # remaining list
        remLst = lst[:i] + lst[i + 1:]

        # Generating all permutations where m is first
        # element
        for p in permutation(remLst):
            l.append([m] + p)
    return l


# Enter a random point to address column in Child table
def enter_random_point(number, data_file):
    for t in range(number):
        temp = "D" + str(t + 2)
        (number_row, y) = random.uniform(-10, 10), random.uniform(-10, 10)
        (number_row, y) = (round(number_row, 2), round(y, 2))
        child_sheet[temp] = str(number_row) + ',' + str(y)

    # Increases the  child table
    tmp = [ws.tables for ws in excel_file.worksheets]
    tables_in_sheet = [{v.name: v} for t in tmp for v in t.values()]

    for table in tables_in_sheet:
        if list(table.keys())[0] == 'Child':
            temp_str = 'A1:G' + str(t + 2)
            table['Child'].ref = temp_str

    excel_file.save(data_file)


# Function to print all the point and all the time to excel
def print_matrix_to_excel():
    # create workbook
    wb = openpyxl.Workbook()
    # get worksheet
    ws = wb.active

    tab = ws.cell(row=1, column=1)
    tab.value = "i\j"
    tab = ws.cell(row=2, column=1)
    tab.value = "(x,y)"
    nColAdd = 1
    length = len(gv.list_of_all_children)
    for j in range(length):
        tab = ws.cell(row=1, column=j + 2 + nColAdd)
        tab.value = str(j + 1)
    tab = ws.cell(row=1, column=length + 2 + nColAdd)
    tab.value = "school"
    for i in range(length + 1):
        tab = ws.cell(row=i + 2, column=1)
        tab.value = str(i + 1)
        tab = ws.cell(row=i + 2, column=2)
        tab.value = str(gv.point_list[i])
        for j in range(length + 1):
            tab = ws.cell(row=i + 2, column=j + 2 + nColAdd)
            tab.value = str(gv.time_matrix[i][j])

    wb.save("matrix.xlsx")


# Creates an excel file for each group
# that contains the travel path, cost and travel time
def print_to_excel_time_cost_per_group(lst, car, cost, time, spath="myPath", sschool="mySchool",
                                       timeOfStart=timedelta(hours=7)):
    # create workbook
    wb = openpyxl.Workbook()
    # get worksheet
    ws = wb.active
    # # change sheet name
    # ws.title = "Group" + str(sheet_number)

    irow = 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "path id"
    tab = ws.cell(row=irow, column=2)
    tab.value = spath
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "car.ID"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(car.ID)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "School"
    tab = ws.cell(row=irow, column=2)
    tab.value = sschool
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "nChildren"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(len(lst))
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "timeOfStart"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(timeOfStart)

    currentTime = timeOfStart
    for t in range(len(lst)):
        # get a pointer for tab in table
        if t > 0:
            iInMatrix = lst[t - 1].id
            jInMatrix = lst[t].id
            temp_time = gv.time_matrix[iInMatrix][jInMatrix]
            currentTime = currentTime + timedelta(minutes=temp_time)
        irow = irow + 1
        tab = ws.cell(row=irow, column=1)
        tab.value = str(t)
        tab = ws.cell(row=irow, column=2)
        tab.value = str(lst[t].id)
        tab = ws.cell(row=irow, column=3)
        tab.value = lst[t].first_name
        tab = ws.cell(row=irow, column=4)
        tab.value = lst[t].lest_name
        tab = ws.cell(row=irow, column=5)
        tab.value = lst[t].address
        tab = ws.cell(row=irow, column=6)
        tab.value = lst[t].contacts
        tab = ws.cell(row=irow, column=7)
        tab.value = str(currentTime)
    iInMatrix = lst[len(lst) - 1].id
    jInMatrix = len(gv.list_of_all_children)
    temp_time = gv.time_matrix[iInMatrix][jInMatrix]
    currentTime = currentTime + timedelta(minutes=temp_time)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "school"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(gv.list_of_school[0].id)
    tab = ws.cell(row=irow, column=3)
    tab.value = gv.list_of_school[0].name
    tab = ws.cell(row=irow, column=4)
    tab.value = sschool
    tab = ws.cell(row=irow, column=5)
    tab.value = gv.list_of_school[0].address
    tab = ws.cell(row=irow, column=6)
    tab.value = gv.list_of_school[0].contacts
    tab = ws.cell(row=irow, column=7)
    tab.value = str(currentTime)

    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "cost"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(cost)
    irow = irow + 1
    tab = ws.cell(row=irow, column=1)
    tab.value = "time"
    tab = ws.cell(row=irow, column=2)
    tab.value = str(time)
    wb.save("Groups" + str(gv.file_number) + ".xlsx")


def calculate_euclidean_dist(start, destination):
    (x_1, y_1) = start
    (x_2, y_2) = destination
    temp_var = (x_1 - x_2) ** 2 + (y_1 - y_2) ** 2
    return math.sqrt(temp_var) * random.uniform(1, 1.5)


# For each group calculate the time and cost
# The function checks all possible path and choosing the shortest path
def calculate_time_cost_per_group(list_of_children, car, school_address):
    total_cost = car.driver_cost
    total_cost_temp_perm = car.driver_cost
    total_time = 0
    total_time_temp_perm = 0
    length = len(list_of_children)
    short_path = []
    flag = 0
    for temp_list_of_children in permutation(list_of_children):
        times = []
        addresses = [temp_list_of_children[0].address]
        for i in range(length - 1):
            (cost, time) = car.calculate_cost_time(temp_list_of_children[i].address,
                                                   temp_list_of_children[i + 1].address)
            total_cost_temp_perm += cost
            total_time_temp_perm += time
            times.append(time)
            addresses.append(temp_list_of_children[i + 1].address)
        # calculate the time and cost for school
        (cost, time) = car.calculate_cost_time(temp_list_of_children[length - 1].address, school_address)
        addresses.append(temp_list_of_children[length - 1].address)  # twice
        addresses.append(school_address)
        times.append(time)
        cost = 0
        if time > car.tMin:
            cost = (time - car.tMin) * car.cost_per_minute
        cost = car.driver_cost + cost
        total_cost_temp_perm = cost  # += cost
        total_time_temp_perm += time

        if total_time_temp_perm < total_time or flag == 0:
            flag = 1
            short_path = temp_list_of_children
            total_time = total_time_temp_perm
            total_cost = total_cost_temp_perm
        total_cost_temp_perm = car.driver_cost
        total_time_temp_perm = 0

    print('group:' + str(gv.file_number))

    print_to_excel_time_cost_per_group(short_path, car, total_cost, total_time)

    list_of_address_in_short_path = []
    for child in short_path:
        (x, y) = child.address.split(',')
        (x, y) = (float(x), float(y))
        list_of_address_in_short_path.append((x, y))

    G = nx.Graph()
    length = len(list_of_address_in_short_path)
    for i in range(length):
        G.add_node(i, pos=list_of_address_in_short_path[i])
    for i in range(length - 1):
        G.add_edge(i, i + 1)
    pos = nx.get_node_attributes(G, 'pos')
    nx.draw(G, pos)
    plt.savefig('G' + str(gv.file_number) + '.png')
    # To delete all the nodes and edges
    plt.clf()


# starting time
start = datetime.now()

# Using openpyxl to writing to excel file
# Give the location of the file
data_file = "Worksheet.xlsx"

# To open Workbook
excel_file = openpyxl.load_workbook(data_file)

# To open sheets
child_sheet = excel_file.worksheets[0]

number_of_children = 21

enter_random_point(number_of_children, data_file)

# get data from excel file
for number_row in range(2, number_of_children + 2):
    ch = Child(child_sheet, number_row)
    ch.id = number_row - 2
    gv.list_of_all_children.insert((number_row - 2), (ch))

# Reading from excel file from Car table
cars_sheet = excel_file.worksheets[2]

# get cars data from data sheet
for number_row in range(2, 5):
    gv.list_of_cars.insert((number_row - 2), (Car(cars_sheet, number_row)))

# Reading from excel file from School table
schools_sheet = excel_file.worksheets[4]

for number_row in range(2, 3):
    gv.list_of_school.insert((number_row - 2), (School(schools_sheet, number_row)))

# Enter the school's address to the matrix
temp_point = gv.list_of_school[0].address
(x1, y1) = temp_point.split(',')
(x1, y1) = (float(x1), float(y1))
gv.point_list.insert(0, (x1, y1))

# Enter the children's address to the list
for number_row in range(0, number_of_children):
    temp_point = gv.list_of_all_children[number_row].address
    (x1, y1) = temp_point.split(',')
    (x1, y1) = (float(x1), float(y1))
    gv.point_list.insert(number_row, (x1, y1))

# Matrix 22X22, this matrix will contain all the time travel from point A (row0) to point B (column0)
gv.time_matrix = np.zeros((number_of_children + 1, number_of_children + 1))

length = len(gv.list_of_all_children)

for i in range(length + 1):
    for j in range(length + 1):
        gv.time_matrix[i, j] = math.ceil(calculate_euclidean_dist(gv.point_list[i], gv.point_list[j]))
print_matrix_to_excel()

# print(gv.point_list)

divide_list_of_children = divide_list_of_children_by_k_means(3)

plt.scatter(*zip(*gv.clusters[0]), color='black')
plt.scatter(*zip(*gv.clusters[1]), color='blue')
plt.scatter(*zip(*gv.clusters[2]), color='red')

plt.savefig('cluster.png')
plt.clf()

length = len(divide_list_of_children)
for i in range(length):
    gv.file_number = i
    calculate_time_cost_per_group(divide_list_of_children[i], gv.list_of_cars[i], gv.list_of_school[0].address)

# end time
end = datetime.now()

# total time taken
print(f"Runtime of the program is {end - start}")
